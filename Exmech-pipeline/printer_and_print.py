import pyautogui
import subprocess
import time
import requests
import json
from openpyxl import load_workbook


class printer:
    def __init__(self, printer_serial):
        self.printer_serial = printer_serial

    # Members
    print_num = 0
    last_batch_serial = None

    # Preform_click
    def preform_click(self, file_name, printer_serial):
        exe_path = "D:/preform/3.34.3.315/PreForm.exe"
        file_path = 'C:/temp/display_stl'
        subprocess.Popen([exe_path, file_path + '/' + file_name])
        print('Preparing to upload print task')
        pyautogui.countdown(20)
        for i in range(10):
            try:
                location = pyautogui.locateOnScreen('print_start.png', confidence=0.95, grayscale=False)
                pyautogui.click(location[0:4], duration=0.5)
                print('Upload successful')
                upload_success = True
                break
            except:
                print('Uploading...')
                pyautogui.countdown(8)
        print('The specified program is not open')
        upload_success = False
        # Close window
        time.sleep(1)
        self.clickicon('not_save.png')
        pyautogui.moveTo(20, 20)
        return upload_success

    # Preform_visit
    def printer_status(self):
        # Get token, initialize
        url = 'https://api.formlabs.com/developer/v1/o/token/'
        header = {'Content-Type': 'application/x-www-form-urlencoded'}
        payload = {
            'grant_type': 'password',
            'client_id': '0Qgg7tKtiLBl9oT5OzDhg4D40Fs3GMmFNo7owcSq',
            'client_secret': 'I9Ydzv0aRmg0HhXrb8FagU1Jw41oowSWWNwSZ7jsht5sM0nadycwVtO78ZqOaP7AWIkcbRQaGHQbNPe5eXeA4uFa1QYDikvSbxxzr2MahwzTTnv8A0pwIusbQj10IZOS',
            'username': 'tiantian',
            'password': '19991225LIFE'
        }
        res = requests.post(url, headers=header, data=payload)
        token_byte = res.content
        token_str = token_byte.decode('utf-8')  # Decode byte to string
        token_json = json.loads(token_str)  # Convert string to json, stored as dictionary

        # Get printer information from the account
        url = 'https://api.formlabs.com/developer/v1/printers/' + self.printer_serial + '/'
        authorization = 'Bearer' + " " + token_json.get('access_token')
        header = {'Authorization': authorization}
        res = requests.get(url=url, headers=header)
        inform = res.content.decode('utf-8')
        json_inform = json.loads(inform)
        print(json_inform)
        # Get device status summary
        is_ready = json_inform.get('printer_status').get('ready_to_print')
        is_last_print_success = json_inform.get('previous_print_run').get('print_run_success')
        is_last_print_harvest = json_inform.get('previous_print_run').get('harvest_status')

        return is_ready, is_last_print_success, is_last_print_harvest


class batch:
    """
    Batch class - experiment batch
    1 batch - 3 specimens
    Members:
    batch_serial (str)
    design_time
    design_feature1
    design_feature2
    design_feature3
    printer_serial
    print_start_time
    is_out_cage
    is_clean
    is_UV
    """
    def __init__(self, batch_serial):
        self.batch_serial = batch_serial
        self.design_time = None
        self.design_feature1 = None
        self.design_feature2 = None
        self.design_feature3 = None
        self.printer_serial = None
        self.print_start_time = None
        self.is_out_cage = False
        self.is_clean = False
        self.is_UV = False


class specimen:
    """
    Specimen class
    Members:
    specimen_serial - specimen number
    design_time
    design_feature1
    printer_serial
    print_start_time
    is_out_cage
    is_clean
    is_UV
    weight
    mechanics_feature:
    shear_strength
    shear_yield_strength
    yield_strength
    stiffness
    toughness
    Methods:
    export_to_excel(filename) - Export to Excel
    """
    def __init__(self, specimen_serial):
        self.specimen_serial = specimen_serial
        self.design_time = None
        self.design_feature1 = None
        self.design_feature2 = None
        self.design_feature3 = None
        self.printer_serial = None
        self.print_start_time = None
        self.is_out_cage = False
        self.is_clean = False
        self.is_UV = False
        self.weight = None
        self.sp_shear_strength = None
        self.sp_shear_yield_strength = None
        self.sp_yield_strength = None
        self.sp_stiffness = None
        self.sp_toughness = None

    def export_to_excel(self, filename, row_num):
        """
        :param filename: Excel path
        :param row_num: Row number to save
        :return:
        """
        wb = load_workbook(filename)
        ws = wb['Sheet1']
        # Get next empty row
        next_row = ws.max_row + 1
        # Write class members into Excel row
        members = [self.specimen_serial, self.design_time, self.design_feature1, self.design_feature2, self.design_feature3, self.printer_serial, self.print_start_time, self.is_out_cage,
                   self.is_clean, self.is_UV, self.weight, self.sp_shear_strength, self.sp_shear_yield_strength, self.sp_yield_strength,
                   self.sp_stiffness, self.sp_toughness]

        for col_num, value in enumerate(members, start=1):
            ws.cell(row=row_num, column=col_num, value=value)

        wb.save(filename)


if __name__ == '__main__':
    pass
