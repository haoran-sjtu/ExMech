import time
import numpy as np
import pandas as pd
from openpyxl import load_workbook
import robot_waypoint
import printer_and_print
import scale
import ultracleanner_pc
import grasp
from grasp import robot
import slider_pc
from dryer_pc import dryer_control
import UTM_client
import model_update

# Robot login
ip = '192.168.20.200'
port = 8899
result = robot.connect(ip, port)
print(f'robot connect is {result}')
# Robot power on
ret = robot.robot_startup()
print("robot_startup ret is {0}".format(ret))
robot.set_collision_class(10)
robot.enable_robot_event()
robot.init_profile()
robot.move_stop()

def station_init():
    robot.set_board_io_status(5, "U_DO_04", 1)  # Set IO4 to 1
    time.sleep(2)
    robot.set_board_io_status(5, "U_DO_04", 0)  # Set IO4 to 0
    time.sleep(1)

    robot.set_board_io_status(5, "U_DO_05", 1)  # Set IO5 to 1
    time.sleep(2)
    robot.set_board_io_status(5, "U_DO_05", 0)  # Set IO5 to 0
    return

def gripper_operation(order):
    """
    Gripper control
    :param order: open, close
    """
    if order == 'open':
        # Open gripper - large
        robot.set_board_io_status(5, "U_DO_00", 1)  # Set IO0 to 1
        robot.set_board_io_status(5, "U_DO_01", 1)  # Set IO1 to 1
    elif order == 'close':
        # Close gripper
        robot.set_board_io_status(5, "U_DO_00", 0)  # Set IO0 to 0
        robot.set_board_io_status(5, "U_DO_01", 0)  # Set IO1 to 0
    elif order == 'mid':
        # Open gripper - medium
        robot.set_board_io_status(5, "U_DO_00", 0)  # Set IO0 to 0
        robot.set_board_io_status(5, "U_DO_01", 1)  # Set IO1 to 1
    elif order == 'minor':
        # Open gripper - small
        robot.set_board_io_status(5, "U_DO_00", 1)  # Set IO0 to 1
        robot.set_board_io_status(5, "U_DO_01", 0)  # Set IO1 to 0
    else:
        print('Gripper command incorrect')
    return

def relative_move(axis, distance):
    """
    Move relative to current position along specified axis
    """
    ik_result = []
    if axis == 'x':
        current_pos = robot.get_current_waypoint()
        current_pos['pos'][0] += distance  # Along X-axis, X increase
        ik_result = robot.inverse_kin(current_pos['joint'], current_pos['pos'], current_pos['ori'])
        robot.move_line(ik_result['joint'])
    elif axis == 'y':
        current_pos = robot.get_current_waypoint()
        current_pos['pos'][1] += distance  # Along Y-axis, Y increase
        ik_result = robot.inverse_kin(current_pos['joint'], current_pos['pos'], current_pos['ori'])
        robot.move_line(ik_result['joint'])
    elif axis == 'z':
        current_pos = robot.get_current_waypoint()
        current_pos['pos'][2] += distance  # Along Z-axis, Z increase
        ik_result = robot.inverse_kin(current_pos['joint'], current_pos['pos'], current_pos['ori'])
        robot.move_line(ik_result['joint'])
    else:
        print('Command incorrect')
    return

def set_speed_level(level):
    """
    Robot arm speed control
    """
    if level == 'slow_speed':
        joint_maxvelc = (0.25, 0.25, 0.25, 0.25, 0.25, 0.25)
        joint_maxacc = (0.8, 0.8, 0.8, 0.8, 0.8, 0.8)
        robot.set_joint_maxacc(joint_maxacc)
        robot.set_joint_maxvelc(joint_maxvelc)
    elif level == 'mid_speed':
        joint_maxvelc = (0.5, 0.5, 0.5, 0.5, 0.5, 0.5)
        joint_maxacc = (0.6, 0.6, 0.6, 0.6, 0.6, 0.6)
        robot.set_joint_maxacc(joint_maxacc)
        robot.set_joint_maxvelc(joint_maxvelc)
    elif level == 'fast_speed':
        joint_maxvelc = (2, 2, 2, 2, 2, 2)
        joint_maxacc = (3, 3, 3, 3, 3, 3)
        robot.set_joint_maxacc(joint_maxacc)
        robot.set_joint_maxvelc(joint_maxvelc)
    else:
        print('Speed adjustment command incorrect')
    return

def flip_specimen():
    """
    Flip specimen
    """
    set_speed_level('mid_speed')
    robot.move_joint(robot_waypoint.Tray.flip_prepare1)
    robot.move_joint(robot_waypoint.Tray.flip_prepare2)
    robot.move_joint(robot_waypoint.Tray.flip1)
    set_speed_level('fast_speed')
    robot.move_joint(robot_waypoint.Tray.flip2)
    time.sleep(1.5)
    robot.move_joint(robot_waypoint.Tray.flip1)
    time.sleep(0.5)
    robot.move_joint(robot_waypoint.Tray.flip2)
    time.sleep(1.5)
    set_speed_level('mid_speed')
    robot.move_joint(robot_waypoint.Tray.flip_prepare1)
    return

def remove_shear_specimen():
    """
    Remove damaged specimen from fixture after shear test
    """
    # Pick
    set_speed_level('mid_speed')
    robot.move_joint(robot_waypoint.ZQ_UTM.UTM_pre_far)
    robot.move_joint(robot_waypoint.ZQ_UTM.UTM_pre)
    gripper_operation('minor')
    set_speed_level('slow_speed')
    robot.move_line(robot_waypoint.ZQ_UTM.remove)  # Move to waypoint
    time.sleep(1)
    gripper_operation('close')
    time.sleep(1)
    robot.move_line(robot_waypoint.ZQ_UTM.UTM_pre)
    set_speed_level('mid_speed')
    robot.move_joint(robot_waypoint.ZQ_UTM.UTM_pre_far)
    # Place specimen
    robot.move_joint(robot_waypoint.Tray.photograph)
    robot.move_joint(robot_waypoint.Tray.shear_recycle)
    time.sleep(0.5)
    gripper_operation('open')
    time.sleep(0.5)
    return

def post_process(op_printer_id):
    # Workstation initialization
    station_init()
    if op_printer_id == 1:
        Formauto = robot_waypoint.slider1
    elif op_printer_id == 2:
        Formauto = robot_waypoint.slider2
    elif op_printer_id == 3:
        Formauto = robot_waypoint.slider3
    else:
        Formauto = None

    # Post process
    current_batch.is_out_cage = True
    gripper_operation('open')
    set_speed_level('mid_speed')
    robot.move_joint(Formauto.slider_pre_far)
    robot.move_joint(Formauto.slider_pre)
    set_speed_level('slow_speed')
    robot.move_line(Formauto.slider_get)
    time.sleep(1)
    gripper_operation('close')
    time.sleep(1)
    robot.move_line(Formauto.slider_pre)
    set_speed_level('mid_speed')
    robot.move_joint(Formauto.slider_pre_far)
    robot.move_joint(robot_waypoint.Ultracleaner.cleaner_pre)
    robot.move_pause()
    gate_state = ultracleanner_pc.get_cleaner_gate()
    ultracleanner_pc.cleaner_gate_control('open', gate_state)
    robot.move_continue()
    robot.move_joint(robot_waypoint.Ultracleaner.cleaner_pre2)
    robot.move_joint(robot_waypoint.Ultracleaner.cleaner_sink_pre)
    set_speed_level('slow_speed')
    robot.move_line(robot_waypoint.Ultracleaner.cleaner_sink_put)
    time.sleep(1)
    gripper_operation('open')
    time.sleep(1)
    robot.move_line(robot_waypoint.Ultracleaner.cleaner_sink_pre)
    set_speed_level('mid_speed')
    robot.move_joint(robot_waypoint.Ultracleaner.cleaner_pre2)
    robot.move_pause()
    gate_state = ultracleanner_pc.get_cleaner_gate()
    ultracleanner_pc.cleaner_gate_control('close', gate_state)
    time.sleep(1.2)
    robot.set_board_io_status(5, "U_DO_03", 1)
    time.sleep(0.5)
    robot.set_board_io_status(5, "U_DO_03", 0)
    time.sleep(900)
    current_batch.is_clean = True
    gate_state = ultracleanner_pc.get_cleaner_gate()
    ultracleanner_pc.cleaner_gate_control('open', gate_state)
    robot.move_continue()
    robot.move_joint(robot_waypoint.Ultracleaner.cleaner_sink_pre)
    set_speed_level('slow_speed')
    robot.move_line(robot_waypoint.Ultracleaner.cleaner_sink_get)
    time.sleep(1)
    gripper_operation('close')
    time.sleep(1)
    robot.move_line(robot_waypoint.Ultracleaner.cleaner_sink_pre)
    set_speed_level('mid_speed')
    robot.move_joint(robot_waypoint.Ultracleaner.cleaner_sink_shake1)
    time.sleep(4)
    robot.move_joint(robot_waypoint.Ultracleaner.cleaner_sink_shake2)
    time.sleep(4)
    robot.move_line(robot_waypoint.Ultracleaner.cleaner_sink_pre)
    robot.move_joint(robot_waypoint.Ultracleaner.cleaner_pre2)
    robot.move_joint(robot_waypoint.Ultracleaner.cleaner_pre)
    robot.move_pause()
    gate_state = ultracleanner_pc.get_cleaner_gate()
    ultracleanner_pc.cleaner_gate_control('close', gate_state)
    dryer_control('out')
    robot.move_continue()
    robot.move_joint(robot_waypoint.Dryer.dryer_pre)
    set_speed_level('slow_speed')
    robot.move_line(robot_waypoint.Dryer.dryer_put)
    time.sleep(1)
    gripper_operation('open')
    time.sleep(1)
    robot.move_line(robot_waypoint.Dryer.dryer_pre)
    set_speed_level('mid_speed')
    robot.move_joint(Formauto.slider_pre_far)
    robot.move_pause()
    dryer_control('run')
    robot.move_continue()
    robot.move_joint(robot_waypoint.Dryer.dryer_pre)
    set_speed_level('slow_speed')
    robot.move_line(robot_waypoint.Dryer.dryer_get)
    time.sleep(1)
    gripper_operation('close')
    time.sleep(1)
    robot.move_line(robot_waypoint.Dryer.dryer_pre)
    set_speed_level('mid_speed')
    robot.move_joint(Formauto.slider_pre_far)
    robot.move_pause()
    dryer_control('in')
    robot.move_continue()

    # Adjust dryer board pose
    robot.move_joint(robot_waypoint.Dryer.dryer_board_prev)
    set_speed_level('slow_speed')
    robot.move_line(robot_waypoint.Dryer.dryer_board_putv)
    time.sleep(1)
    gripper_operation('open')
    time.sleep(1)
    robot.move_line(robot_waypoint.Dryer.dryer_board_prev)
    set_speed_level('mid_speed')
    robot.move_joint(robot_waypoint.Dryer.dryer_board_preh_pre)
    robot.move_joint(robot_waypoint.Dryer.dryer_board_preh)
    set_speed_level('slow_speed')
    robot.move_line(robot_waypoint.Dryer.dryer_board_geth)
    time.sleep(1)
    gripper_operation('close')
    time.sleep(1)
    robot.move_line(robot_waypoint.Dryer.dryer_board_preh)
    # Flip
    flip_specimen()
    robot.move_joint(robot_waypoint.Dryer.dryer_board_preh)
    set_speed_level('slow_speed')
    robot.move_line(robot_waypoint.Dryer.dryer_board_puth)
    time.sleep(1)
    gripper_operation('open')
    time.sleep(1)
    robot.move_line(robot_waypoint.Dryer.dryer_board_preh)
    # Vertical pick
    set_speed_level('mid_speed')
    robot.move_joint(robot_waypoint.Dryer.dryer_board_prev)
    set_speed_level('slow_speed')
    robot.move_line(robot_waypoint.Dryer.dryer_board_getv)
    time.sleep(1)
    gripper_operation('close')
    time.sleep(1)
    robot.move_line(robot_waypoint.Dryer.dryer_board_prev)
    set_speed_level('mid_speed')
    robot.move_joint(Formauto.slider_pre_far)
    robot.move_joint(Formauto.slider_pre)
    set_speed_level('slow_speed')
    robot.move_line(Formauto.slider_put)
    time.sleep(1)
    gripper_operation('open')
    time.sleep(1)
    robot.move_line(Formauto.slider_pre)
    set_speed_level('mid_speed')
    robot.move_joint(Formauto.slider_pre_far)

def get_weight():
    # Balance weighing
    set_speed_level('mid_speed')
    robot.move_joint(robot_waypoint.Balance.balance_pre)
    robot.move_line(robot_waypoint.Balance.balance_put)
    scale.scale_oprate('clear')
    time.sleep(2)
    gripper_operation('minor')
    time.sleep(1)
    set_speed_level('slow_speed')
    relative_move('z', 0.05)
    time.sleep(3)
    weight = scale.scale_oprate('read')
    gripper_operation('mid')
    set_speed_level('mid_speed')
    robot.move_line(robot_waypoint.Balance.balance_get)
    time.sleep(0.8)
    gripper_operation('close')
    time.sleep(0.8)
    robot.move_line(robot_waypoint.Balance.balance_pre)
    return weight

def compress_test(specimen_serial):
    # Compression test
    set_speed_level('mid_speed')
    robot.move_joint(robot_waypoint.UTM.UTM_pre_far)
    robot.move_line(robot_waypoint.UTM.UTM_pre)
    set_speed_level('slow_speed')
    robot.move_line(robot_waypoint.UTM.fixture)
    time.sleep(1)
    gripper_operation('minor')
    time.sleep(1)
    robot.move_line(robot_waypoint.UTM.UTM_pre)
    set_speed_level('mid_speed')
    robot.move_line(robot_waypoint.UTM.UTM_pre_far)
    robot.move_joint(robot_waypoint.Tray.photograph)
    print('Specimen in place, starting compression test...')
    robot.move_pause()
    # Execute test
    mechanics_data = UTM_client.UTM_control('execute_compress_test', specimen_serial)
    print(f'Received mechanics_features: {mechanics_data}')
    # Clear specimen
    robot.set_board_io_status(5, "U_DO_04", 1)
    time.sleep(1.5)
    robot.set_board_io_status(5, "U_DO_04", 0)
    print('Tray cleaning executed')
    robot.move_continue()
    return mechanics_data

def shear_test(specimen_serial):
    # Shear test
    set_speed_level('mid_speed')
    robot.move_joint(robot_waypoint.Tray.photograph)
    robot.move_joint(robot_waypoint.ZQ_UTM.UTM_pre_far)
    robot.move_joint(robot_waypoint.ZQ_UTM.UTM_pre)
    set_speed_level('slow_speed')
    robot.move_line(robot_waypoint.ZQ_UTM.fixture)
    time.sleep(1)
    gripper_operation('minor')
    time.sleep(1)
    # Centering
    robot.set_board_io_status(5, "U_DO_05", 1)
    time.sleep(1)
    robot.set_board_io_status(5, "U_DO_05", 0)
    time.sleep(0.5)
    robot.set_board_io_status(5, "U_DO_05", 1)
    time.sleep(1)
    robot.set_board_io_status(5, "U_DO_05", 0)
    time.sleep(0.5)
    robot.move_line(robot_waypoint.ZQ_UTM.UTM_pre)
    set_speed_level('mid_speed')
    robot.move_joint(robot_waypoint.ZQ_UTM.monitor)
    robot.move_pause()
    print('Specimen in place, starting shear test...')
    # Execute test
    mechanics_data = UTM_client.UTM_control('execute_shear_test', specimen_serial)
    print(f'Received mechanics_features: {mechanics_data}')
    robot.move_continue()
    return mechanics_data

# Working flow
# Initialization
printer_1 = printer_and_print.printer('ExpertMackerel')
printer_2 = printer_and_print.printer('FamousOriole')
printer_3 = printer_and_print.printer('SportyStork')
printer = [printer_1, printer_2, printer_3]
task_num = 0
task_que = list(range(1, 11))

while True:
    op_printer = None
    op_printer_id = 0
    print('Preparing to check printers...')
    time.sleep(50)
    for i in range(1, 2):
        is_ready, *_ = printer[i - 1].printer_status()
        if is_ready == 'READY_TO_PRINT_NEEDS_CONFIRMATION':
            print('Build platform needs manual cleaning')
            break
        elif is_ready == 'READY_TO_PRINT_NOT_READY':
            print(f'printer{i} is working')
        elif is_ready == 'READY_TO_PRINT_READY':
            print(f'printer{i} is idle')
            op_printer_id = i
            op_printer = printer[i-1]
            break
        else:
            print(f'Printer{i} returns status: {is_ready}')

    if len(task_que) == 0:
        print('Task queue is empty')
    elif op_printer is None:
        print('No idle printer currently, will check again in 5 minutes')
        time.sleep(300)
    else:
        current_task_id = task_que[0]
        if current_task_id % 2 == 1:
            batch_serial = str(round((current_task_id+1)/2))+'_a'
        else:  # Even number
            batch_serial = str(round(current_task_id/2))+'_b'
        print(f'Batch serial to be printed is {batch_serial}, printer to execute printing is {op_printer.printer_serial}')

        filename = f'{batch_serial}.stl'
        op_printer.preform_click(filename, op_printer_id)
        op_printer.print_num += 1
        current_batch = printer_and_print.batch(op_printer.last_batch_serial)
        current_batch.printer_serial = op_printer.printer_serial
        task_que.pop(0)
        line = int(np.ceil(current_task_id / 2))
        df = pd.read_excel('design.xlsx')
        design_mat = df.to_numpy()
        current_batch.design_time = design_mat[line - 1, 1]
        current_batch.design_feature1 = design_mat[line - 1, 2]
        current_batch.design_feature2 = design_mat[line - 1, 3]
        current_batch.design_feature3 = design_mat[line - 1, 4]
        op_printer.last_batch_serial = batch_serial

        # Check if printer is printing for the first time
        if op_printer.print_num <= 1:
            print('This printer is executing its first task, no specimens need post-processing')
        else:
            print('Preparing to execute post-processing...')
            slider_pc.slider_control('out', op_printer_id)
            robot.move_continue()
            post_process(op_printer_id)
            robot.move_pause()
            slider_pc.slider_control('in', op_printer_id)
            print(f'Post-processing completed, batch {current_batch.batch_serial} specimens ready for mechanical testing...')
            robot.move_continue()

            for n in range(1, 4):
                set_speed_level('mid_speed')
                grasp.grasp_main()
                specimen_serial = f'{current_batch.batch_serial}_{n}'
                specimen = printer_and_print.specimen(specimen_serial)
                print(f'Picked {specimen_serial}')
                specimen.design_feature1 = current_batch.design_feature1
                specimen.design_feature2 = current_batch.design_feature2
                specimen.design_feature3 = current_batch.design_feature3
                specimen.printer_serial = current_batch.printer_serial
                specimen.design_time = current_batch.design_time
                specimen.is_clean = current_batch.is_clean
                specimen.is_out_cage = current_batch.is_out_cage
                specimen.print_start_time = current_batch.print_start_time
                weight = get_weight()
                specimen.weight = weight

                if specimen_serial[2] == 'a':
                    # Shear test
                    mechanics_shear = shear_test(specimen_serial)
                    # Assign specimen members (mechanics_shear)
                    specimen.sp_shear_strength = mechanics_shear['shear_strength']/specimen.weight
                    specimen.sp_shear_yield_strength = mechanics_shear['shear_yield_strength']/specimen.weight
                    # Remove material
                    remove_shear_specimen()
                    print('Material removed')
                    row_num = 6*(int(specimen_serial[0])-1)+n+1

                elif specimen_serial[2] == 'b':
                    # Compression test
                    mechanics_compress = compress_test(specimen_serial)
                    # Assign specimen members
                    specimen.sp_yield_strength = mechanics_compress['yield_strength']/specimen.weight
                    specimen.sp_stiffness = mechanics_compress['stiffness']/specimen.weight
                    specimen.sp_toughness = mechanics_compress['toughness']/specimen.weight
                    row_num = 6*int(specimen_serial[0])-3+n+1
                else:
                    print('Unable to identify specimen type')
                    row_num = None
                # Export to specimen_raw.xlsx
                specimen.export_to_excel('specimen_raw.xlsx', row_num)

            sp_raw = pd.read_excel('specimen_raw.xlsx')
            specimen_mat = sp_raw.to_numpy()
            specimen_xlsx_row = int(current_batch.batch_serial[0])+1
            if current_batch.batch_serial[2] == 'a':
                mat_shear = specimen_mat[-3:, 11:13]
                np.mean(mat_shear[:, 0])
                array_shear = [np.mean(mat_shear[:,0]), np.std(mat_shear[:,0],ddof=1), np.mean(mat_shear[:, 1]),
                               np.std(mat_shear[:,1],ddof=1)]
                wb = load_workbook('specimen.xlsx')
                ws = wb['Sheet1']
                for i, value in enumerate(array_shear):
                    col = 11+i
                    ws.cell(row=specimen_xlsx_row, column=col, value=value)
                mat_design = specimen_mat[-1, 0:10]
                mat_design[0] = mat_design[0][0]
                for i, value in enumerate(mat_design):
                    col = 1 + i
                    ws.cell(row=specimen_xlsx_row, column=col, value=value)
                wb.save('specimen.xlsx')
                print("Shear data written to Excel file")
            elif current_batch.batch_serial[2] == 'b':
                mat_compress = specimen_mat[-3:, 13:16]  # 3*3
                array_compress = [np.mean(mat_compress[:, 0]), np.std(mat_compress[:, 0], ddof=1), np.mean(mat_compress[:, 1]),
                                  np.std(mat_compress[:, 1], ddof=1), np.mean(mat_compress[:, 2]), np.std(mat_compress[:, 2], ddof=1)]
                wb = load_workbook('specimen.xlsx')
                ws = wb['Sheet1']
                for i, value in enumerate(array_compress):
                    col = 15 + i
                    ws.cell(row=specimen_xlsx_row, column=col, value=value)
                wb.save('specimen.xlsx')
                print("Compression data written to Excel file")

            task_num += 1
            print(f'update task_num: {task_num}')
            # Robot returns to initial position
            robot.move_joint(robot_waypoint.Tray.photograph)
            robot.move_joint(robot_waypoint.slider1.slider_pre_far)
            robot.move_stop()  # Robot stops

            is_new_design = model_update.update_design()

            if is_new_design is True:
                print('Design parameters updated, preparing to generate specimen 3D model')
                design = pd.read_excel('design.xlsx')
                design_mat = design.to_numpy()
                design_feature = {'batch_serial': str(design_mat[-1, 0]), 'd1': design_mat[-1, 2], 'd2': design_mat[-1, 3], 'd3': design_mat[-1, 4]}


            # Update task_que
            task_que.append(max(task_que)+1)